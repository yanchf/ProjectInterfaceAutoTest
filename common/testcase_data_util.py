import json
from pathlib import Path
import openpyxl
import requests
from config.config_file import xlsx_format, env_xlsx_format
from run_test import arg_env
from common.loggin_util import LoggingUtil

# 初始化日志工具，记录程序运行状态
logging_util_temp = LoggingUtil(__name__)
logging_util = logging_util_temp.init_logger()


# 获取符合规则的excel文件绝对路径s
def get_testcase_file(path=r'testcase_data/excel'):
    file_group = []
    p = Path(path)
    file_list = list(p.rglob("testcase*.xlsx"))

    for file_name in file_list:
        if "模板" not in file_name.stem:
            file_group.append(file_name)
    logging_util.info(file_group)
    return file_group


# 获取token
def _init_tokens(file_list: list = get_testcase_file()):
    # 存放测试环境登陆数据
    testcase_token = dict()
    # 存放正式环境登陆数据
    testcase_token_real = dict()

    for work_xlsx in file_list:
        wb = openpyxl.load_workbook(work_xlsx)
        if ("init_token" not in wb.sheetnames) and ("init_token_real" not in wb.sheetnames):
            raise TypeError("xlsx格式错误")
        else:
            # 获取测试环境登录相关接口
            sheet = wb["init_token"]
            # 获取正式环境相关接口
            sheet_real = wb["init_token_real"]

            testcase_token[work_xlsx.stem] = list()
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=sheet.min_column,
                                       max_col=sheet.max_column, values_only=True):
                testcase_token[work_xlsx.stem].append(dict(zip(xlsx_format, row)))

            testcase_token_real[work_xlsx.stem] = list()
            for row in sheet_real.iter_rows(min_row=2, max_row=sheet.max_row, min_col=sheet.min_column,
                                       max_col=sheet.max_column, values_only=True):
                testcase_token_real[work_xlsx.stem].append(dict(zip(xlsx_format, row)))

    return {"test_env": testcase_token, "real_env": testcase_token_real}


# 获取测试接口集合
def get_testcases(file_list: list = get_testcase_file()):
    # 存放所有文件的测试接口数据
    testcase_data = list()

    for work_xlsx in file_list:
        wb = openpyxl.load_workbook(work_xlsx)
        if "testcase_data" not in wb.sheetnames:
            raise TypeError("xlsx格式错误")
        else:
            sheet = wb["testcase_data"]
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=sheet.min_column,
                                       max_col=sheet.max_column, values_only=True):
                testcase_data_temp = dict(zip(xlsx_format, row))
                testcase_data_temp["project"] = work_xlsx.stem
                testcase_data.append(testcase_data_temp)

    logging_util.info(testcase_data)
    return testcase_data


def _init_env(path=r'testcase_data/env/env.xlsx'):
    env_data = dict()
    wb = openpyxl.load_workbook(path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=sheet.min_column, max_col=sheet.max_column,
                               values_only=True):
        env_data[row[0]] = {'test_env': row[1], 'real_env': row[2]}

    return env_data


# 获取excel->init_token下所有系统token
def get_tokens():
    """
    获取token的重要功能，不同项目可能获取账号token的规则不同，
    无法兼容其他获取token方式，如有其他获取token规则，需要修改这个函数
    :return: dict
    """
    # 存放所有项目对应的token
    token_data = dict()
    # 请求头字典
    headers = None

    # 获取token集合
    # 按照对应测试环境获取登录相关的接口字典
    token_dict = _init_tokens()[arg_env]
    # 获取系统正式环境/测试环境
    env_dict = _init_env()

    # 抽取
    tokens_list = list(token_dict.keys())

    # 需要获取不同项目的token，并以文件名为键名组成一个字典
    # 超级码3.0系统一般是先登录，然后再绑定企业再绑定账号，如果有其他系统特殊规则修改for循环内代码
    for token_project in tokens_list:
        for api in token_dict[token_project]:
            # 其实这个流程是有问题的，如果excel里login/org/sys没有按照顺序填写就会报错
            if api["name"] == "login":
                # 登录，获取初识token
                logging_util.info(env_dict[token_project][arg_env] + api["api"])
                logging_util.info(json.loads(api['json_data']))

                response = requests.request(method=api["method"], url=env_dict[token_project][arg_env] + api["api"],
                                            json=json.loads(api['json_data']))
                try:
                    logging_util.info(response.json())
                    super_token = response.json()["results"]["token"]

                    headers = {"super-token": super_token}
                except KeyError as e:
                    raise KeyError("login获取token失败")

            # 绑定企业
            elif (api["name"] == "org") and (headers is not None):
                logging_util.info(env_dict[token_project][arg_env] + api["api"])
                logging_util.info(json.loads(api['json_data']))

                response = requests.request(method=api["method"], url=env_dict[token_project][arg_env] + api["api"],
                                            json=json.loads(api['json_data']), headers=headers)
            # 绑定系统
            elif (api["name"] == "sys") and headers is not None:
                logging_util.info(env_dict[token_project][arg_env] + api["api"])
                logging_util.info(json.loads(api['json_data']))

                # 这个地方我也不知道为啥要设置一个isAdmin=false
                headers.update({'isAdmin': 'false'})

                response = requests.request(method=api["method"], url=env_dict[token_project][arg_env] + api["api"],
                                            json=json.loads(api['json_data']), headers=headers)

                token_data[token_project] = response.request.headers["super-token"]

    logging_util.info(token_data)
    return token_data


# 获取不同操作系统的测试环境数据
def get_env() -> dict:
    return _init_env()
